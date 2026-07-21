#!/usr/bin/env python3
"""
Gera os arquivos de dados (data/videos.json e data/playlists.json) do
catálogo de vídeos da IPA Limeira a partir da planilha exportada.

Uso:
    python3 gerar_catalogo.py planilha.xlsx
    python3 gerar_catalogo.py planilha.xlsx --saida ./minha_pasta

O script NÃO toca em index.html / style.css / app.js — ele só gera os
dados. Assim, rodar de novo (planilha atualizada) nunca sobrescreve o
front-end, só atualiza o conteúdo.
"""

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Este script precisa do pacote openpyxl. Instale com:\n"
              "  pip install openpyxl --break-system-packages")

SHEET_NAME = "Catálogo"

# Ordem canônica dos 66 livros, usada para o filtro Antigo/Novo Testamento
# e para ordenar gráficos por ordem bíblica.
BOOK_ORDER = [
    "Gênesis", "Êxodo", "Levítico", "Números", "Deuteronômio", "Josué",
    "Juízes", "Rute", "1 Samuel", "2 Samuel", "1 Reis", "2 Reis",
    "1 Crônicas", "2 Crônicas", "Esdras", "Neemias", "Ester", "Jó",
    "Salmos", "Provérbios", "Eclesiastes", "Cantares", "Isaías",
    "Jeremias", "Lamentações", "Ezequiel", "Daniel", "Oséias", "Joel",
    "Amós", "Obadias", "Jonas", "Miquéias", "Naum", "Habacuque",
    "Sofonias", "Ageu", "Zacarias", "Malaquias",
    "Mateus", "Marcos", "Lucas", "João", "Atos", "Romanos",
    "1 Coríntios", "2 Coríntios", "Gálatas", "Efésios", "Filipenses",
    "Colossenses", "1 Tessalonicenses", "2 Tessalonicenses",
    "1 Timóteo", "2 Timóteo", "Tito", "Filémon", "Hebreus", "Tiago",
    "1 Pedro", "2 Pedro", "1 João", "2 João", "3 João", "Judas",
    "Apocalipse",
]
OT_COUNT = 39
BOOK_INDEX = {b: i for i, b in enumerate(BOOK_ORDER)}

# Correções pontuais observadas na planilha (espaçamento, digitação,
# ou vazamento de referência de versículo para dentro do campo Livro).
BOOK_FIXES = {
    "Isaías": "Isaías",
    "1Coríntios": "1 Coríntios",
    "2 Coríntios": "2 Coríntios",
    "1 Tessalonicensses": "1 Tessalonicenses",
    "Oseias": "Oséias",
}


def clean(value):
    """Normaliza string: tira espaço nas pontas e trata placeholders vazios."""
    if value is None:
        return None
    s = str(value).strip()
    if s in ("", "---", "--", "-"):
        return None
    return s


def fix_book(value):
    v = clean(value)
    if v is None:
        return None
    v = re.sub(r"\s+", " ", v).strip()
    if v in BOOK_FIXES:
        v = BOOK_FIXES[v]
    if v in BOOK_INDEX:
        return v
    # Campo com lixo colado (ex.: "Atos 7.58 a") -> acha o nome do livro
    # como prefixo, do mais longo pro mais curto pra não confundir
    # "1 Reis" com "1 Re..." etc.
    for book in sorted(BOOK_ORDER, key=len, reverse=True):
        if v.startswith(book):
            return book
    return v  # devolve como veio; melhor mostrar do que descartar


def extract_video_id(url):
    if not url:
        return None
    m = re.search(r"(?:youtu\.be/|v=|embed/)([A-Za-z0-9_-]{11})", url)
    return m.group(1) if m else None


def parse_duration_seconds(s):
    if not s:
        return None
    parts = str(s).split(":")
    try:
        parts = [int(p) for p in parts]
    except ValueError:
        return None
    while len(parts) < 3:
        parts.insert(0, 0)
    h, m, sec = parts[-3:]
    return h * 3600 + m * 60 + sec


def load_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        print(f"Aviso: aba '{SHEET_NAME}' não encontrada, usando a primeira aba "
              f"('{wb.sheetnames[0]}').", file=sys.stderr)
        ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return [r for r in rows if r and r[0]]


def build_video(row):
    (yt_id, url, data, texto_base, categoria, preletor, titulo, duracao,
     ano, livro, playlist_id, playlist_name, *_rest) = (list(row) + [None] * 12)[:12]

    vid = clean(yt_id) or extract_video_id(url)
    if not vid:
        return None

    data_str = None
    if isinstance(data, datetime):
        data_str = data.strftime("%Y-%m-%d")
    elif clean(data):
        data_str = clean(data)

    livro_norm = fix_book(livro)
    testamento = None
    if livro_norm in BOOK_INDEX:
        testamento = "AT" if BOOK_INDEX[livro_norm] < OT_COUNT else "NT"

    playlist_id_clean = clean(playlist_id)
    playlist_name_clean = clean(playlist_name)

    return {
        "video_id": vid,
        "url": f"https://youtu.be/{vid}",
        "thumbnail_url": f"https://img.youtube.com/vi/{vid}/hqdefault.jpg",
        "embed_url": f"https://www.youtube.com/embed/{vid}",
        "data": data_str,
        "ano": int(ano) if ano else (int(data_str[:4]) if data_str else None),
        "texto_base": clean(texto_base),
        "categoria": clean(categoria),
        "preletor": clean(preletor),
        "titulo": clean(titulo),
        "duracao": clean(duracao),
        "duracao_seg": parse_duration_seconds(duracao),
        "livro": livro_norm,
        "testamento": testamento,
        "playlist_id": playlist_id_clean,
        "playlist_name": playlist_name_clean,
        "playlist_url": (
            f"https://www.youtube.com/playlist?list={playlist_id_clean}"
            if playlist_id_clean else None
        ),
    }


def build_playlists(videos):
    agg = {}
    for v in videos:
        pid = v["playlist_id"]
        if not pid:
            continue
        entry = agg.setdefault(pid, {
            "playlist_id": pid,
            "playlist_name": v["playlist_name"] or pid,
            "playlist_url": v["playlist_url"],
            "count": 0,
            "primeira_data": None,
            "ultima_data": None,
        })
        entry["count"] += 1
        d = v["data"]
        if d:
            if not entry["primeira_data"] or d < entry["primeira_data"]:
                entry["primeira_data"] = d
            if not entry["ultima_data"] or d > entry["ultima_data"]:
                entry["ultima_data"] = d
    return sorted(agg.values(), key=lambda p: p["count"], reverse=True)


def main():
    parser = argparse.ArgumentParser(description="Gera os dados do catálogo IPA Limeira")
    parser.add_argument("planilha", help="Caminho para o arquivo .xlsx exportado")
    parser.add_argument("--saida", default=".", help="Pasta de saída (padrão: pasta atual)")
    args = parser.parse_args()

    xlsx_path = Path(args.planilha)
    if not xlsx_path.exists():
        sys.exit(f"Arquivo não encontrado: {xlsx_path}")

    saida = Path(args.saida) / "data"
    saida.mkdir(parents=True, exist_ok=True)

    rows = load_rows(xlsx_path)
    videos = []
    ignorados = 0
    for r in rows:
        v = build_video(r)
        if v:
            videos.append(v)
        else:
            ignorados += 1

    videos.sort(key=lambda v: v["data"] or "", reverse=True)
    playlists = build_playlists(videos)

    with open(saida / "videos.json", "w", encoding="utf-8") as f:
        json.dump(videos, f, ensure_ascii=False, separators=(",", ":"))

    with open(saida / "playlists.json", "w", encoding="utf-8") as f:
        json.dump(playlists, f, ensure_ascii=False, separators=(",", ":"))

    print(f"OK: {len(videos)} vídeos processados ({ignorados} linha(s) ignorada(s) sem ID).")
    print(f"OK: {len(playlists)} playlists identificadas.")
    sem_livro = sum(1 for v in videos if not v["livro"])
    sem_preletor = sum(1 for v in videos if not v["preletor"])
    print(f"Aviso: {sem_livro} vídeo(s) sem livro identificado, {sem_preletor} sem preletor identificado.")
    print(f"Arquivos gravados em: {saida.resolve()}")


if __name__ == "__main__":
    main()
